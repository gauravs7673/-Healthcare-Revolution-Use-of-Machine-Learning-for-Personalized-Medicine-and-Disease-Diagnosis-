from tkinter import *
from tkinter import messagebox
import sqlite3
import numpy as np
import pandas as pd
import openpyxl
from openpyxl import workbook, load_workbook
from sklearn.ensemble import RandomForestClassifier
from sklearn.tree import DecisionTreeClassifier
from sklearn.neighbors import KNeighborsClassifier
from sklearn.model_selection import train_test_split
from sklearn.metrics import precision_score, recall_score, f1_score
from tkinter import StringVar, Checkbutton
import matplotlib.pyplot as plt
import os
import os.path
import warnings

global var1,var2,var3,var4,var5
# Suppress the warning
warnings.filterwarnings("ignore", category=UserWarning)

# Function to calculate precision, recall, and F1 score
def calculate_metrics(true_labels, predicted_labels):
    precision = precision_score(true_labels, predicted_labels, average='weighted')
    recall = recall_score(true_labels, predicted_labels, average='weighted')
    f1 = f1_score(true_labels, predicted_labels, average='weighted')
    return precision, recall, f1

###
# # Function to calculate and store metrics for each algorithm
# def calculate_and_store_metrics(algorithm_name, true_labels, predicted_labels):
#     accuracy, precision, recall, f1 = calculate_metrics(true_labels, predicted_labels)
#     accuracy_scores.append(accuracy)
#     precision_scores.append(precision)
#     recall_scores.append(recall)
#     f1_scores.append(f1)
#     print(f"{algorithm_name} Metrics:")
#     print("Accuracy:", accuracy)
#     print("Precision:", precision)
#     print("Recall:", recall)
#     print("F1 Score:", f1)

# # Define lists to store metrics for each algorithm
# algorithms = ['Naive Bayes', 'SVM', 'Random Forest', 'Decision Tree', 'KNN']
# accuracy_scores = []
# precision_scores = []
# recall_scores = []
# f1_scores = []
# ####


# def enter_data():
#     # Retrieve symptom values from the input fields
    
#     symptom1_val = Symptom1.get()
#     symptom2_val = Symptom2.get()
#     symptom3_val = Symptom3.get()
#     symptom4_val = Symptom4.get()
#     symptom5_val = Symptom5.get()

#     get_disease_names = ['Fungal infection','Allergy','GERD','Chronic cholestasis','Drug Reaction',
#         'Peptic ulcer diseae','AIDS','Diabetes','Gastroenteritis','Bronchial Asthma','Hypertension',
#         ' Migraine','Cervical spondylosis',
#         'Paralysis (brain hemorrhage)','Jaundice','Malaria','Chicken pox','Dengue','Typhoid','hepatitis A',
#         'Hepatitis B','Hepatitis C','Hepatitis D','Hepatitis E','Alcoholic hepatitis','Tuberculosis',
#         'Common Cold','Pneumonia','Dimorphic hemmorhoids(piles)',
#         'Heartattack','Varicoseveins','Hypothyroidism','Hyperthyroidism','Hypoglycemia','Osteoarthristis',
#         'Arthritis','(vertigo) Paroymsal  Positional Vertigo','Acne','Urinary tract infection','Psoriasis',
#         'Impetigo']

#     diseases = get_disease_names(['Symptom1_val', 'Symptom2_val', 'Symptom3_val', 'Symptom4_val', 'Symptom5_val'])

#     # Display diseases in console 
#     print("Diseases:", diseases)

#     # Display symptoms in console 
#     print("Symptom 1:", symptom1_val)
#     print("Symptom 2:", symptom2_val)
#     print("Symptom 3:", symptom3_val)
#     print("Symptom 4:", symptom4_val)
#     print("Symptom 5:", symptom5_val)
        
#     # Define the file path for the Excel file
#     filepath = "D:\HealthCare Software Project\healthcare.csv"

#     # If the file doesn't exist, create it and add a header row
#     if not os.path.exists(filepath):
#         workbook = openpyxl.Workbook()
#         sheet = workbook.active
#         heading = ["Symptom1", "Symptom2", "Symptom3", "Symptom4", "Symptom5"]
#         sheet.append(heading)
#         workbook.save(filepath)

    # # Open the existing Excel file and append the symptom values
    # workbook = openpyxl.load_workbook(filepath)
    # sheet = workbook.active
    # sheet.append([symptom1_val, symptom2_val, symptom3_val, symptom4_val, symptom5_val])
    # workbook.save(filepath)




# Path to your existing CSV file
csv_file_path = 'healthcaredata.csv'




l1=['itching','skin_rash','nodal_skin_eruptions','continuous_sneezing','shivering','chills','joint_pain',
    'stomach_pain','acidity','ulcers_on_tongue','muscle_wasting','vomiting','burning_micturition','spotting_ urination','fatigue',
    'weight_gain','anxiety','cold_hands_and_feets','mood_swings','weight_loss','restlessness','lethargy','patches_in_throat',
    'irregular_sugar_level','cough','high_fever','sunken_eyes','breathlessness','sweating','dehydration','indigestion',
    'headache','yellowish_skin','dark_urine','nausea','loss_of_appetite','pain_behind_the_eyes','back_pain','constipation',
    'abdominal_pain','diarrhoea','mild_fever','yellow_urine','yellowing_of_eyes','acute_liver_failure','fluid_overload',
    'swelling_of_stomach','swelled_lymph_nodes','malaise','blurred_and_distorted_vision','phlegm','throat_irritation',
    'redness_of_eyes','sinus_pressure','runny_nose','congestion','chest_pain','weakness_in_limbs','fast_heart_rate',
    'pain_during_bowel_movements','pain_in_anal_region','bloody_stool','irritation_in_anus','neck_pain','dizziness','cramps',
    'bruising','obesity','swollen_legs','swollen_blood_vessels','puffy_face_and_eyes','enlarged_thyroid','brittle_nails',
    'swollen_extremeties','excessive_hunger','extra_marital_contacts','drying_and_tingling_lips','slurred_speech','knee_pain','hip_joint_pain',
    'muscle_weakness','stiff_neck','swelling_joints','movement_stiffness','spinning_movements','loss_of_balance','unsteadiness','weakness_of_one_body_side',
    'loss_of_smell','bladder_discomfort','foul_smell_of urine','continuous_feel_of_urine','passage_of_gases','internal_itching','toxic_look_(typhos)',
    'depression','irritability','muscle_pain','altered_sensorium','red_spots_over_body','belly_pain','abnormal_menstruation','dischromic _patches',
    'watering_from_eyes','increased_appetite','polyuria','family_history','mucoid_sputum','rusty_sputum','lack_of_concentration','visual_disturbances',
    'receiving_blood_transfusion','receiving_unsterile_injections','coma','stomach_bleeding','distention_of_abdomen','history_of_alcohol_consumption',
    'fluid_overload','blood_in_sputum','prominent_veins_on_calf','palpitations','painful_walking','pus_filled_pimples','blackheads','scurring','skin_peeling',
    'silver_like_dusting','small_dents_in_nails','inflammatory_nails','blister','red_sore_around_nose','yellow_crust_ooze']


disease=['Fungal infection','Allergy','GERD','Chronic cholestasis','Drug Reaction',
        'Peptic ulcer diseae','AIDS','Diabetes','Gastroenteritis','Bronchial Asthma','Hypertension',
        ' Migraine','Cervical spondylosis',
        'Paralysis (brain hemorrhage)','Jaundice','Malaria','Chicken pox','Dengue','Typhoid','hepatitis A',
'Hepatitis B','Hepatitis C','Hepatitis D','Hepatitis E','Alcoholic hepatitis','Tuberculosis',
'Common Cold','Pneumonia','Dimorphic hemmorhoids(piles)',
'Heartattack','Varicoseveins','Hypothyroidism','Hyperthyroidism','Hypoglycemia','Osteoarthristis',
'Arthritis','(vertigo) Paroymsal  Positional Vertigo','Acne','Urinary tract infection','Psoriasis',
'Impetigo']

l2=[]
for x in range(0,len(l1)):
    l2.append(0)

# TESTING DATA
tr=pd.read_csv("Testing.csv")

tr.replace({'prognosis':{'Fungal infection':0,'Allergy':1,'GERD':2,'Chronic cholestasis':3,'Drug Reaction':4,
'Peptic ulcer diseae':5,'AIDS':6,'Diabetes ':7,'Gastroenteritis':8,'Bronchial Asthma':9,'Hypertension ':10,
'Migraine':11,'Cervical spondylosis':12,
'Paralysis (brain hemorrhage)':13,'Jaundice':14,'Malaria':15,'Chicken pox':16,'Dengue':17,'Typhoid':18,'hepatitis A':19,
'Hepatitis B':20,'Hepatitis C':21,'Hepatitis D':22,'Hepatitis E':23,'Alcoholic hepatitis':24,'Tuberculosis':25,
'Common Cold':26,'Pneumonia':27,'Dimorphic hemmorhoids(piles)':28,'Heart attack':29,'Varicose veins':30,'Hypothyroidism':31,
'Hyperthyroidism':32,'Hypoglycemia':33,'Osteoarthristis':34,'Arthritis':35,
'(vertigo) Paroymsal  Positional Vertigo':36,'Acne':37,'Urinary tract infection':38,'Psoriasis':39,
'Impetigo':40}},inplace=True)

X_test= tr[l1]
y_test = tr[["prognosis"]]
np.ravel(y_test)

# TRAINING DATA
df=pd.read_csv("Training.csv")

df.replace({'prognosis':{'Fungal infection':0,'Allergy':1,'GERD':2,'Chronic cholestasis':3,'Drug Reaction':4,
'Peptic ulcer diseae':5,'AIDS':6,'Diabetes ':7,'Gastroenteritis':8,'Bronchial Asthma':9,'Hypertension ':10,
'Migraine':11,'Cervical spondylosis':12,
'Paralysis (brain hemorrhage)':13,'Jaundice':14,'Malaria':15,'Chicken pox':16,'Dengue':17,'Typhoid':18,'hepatitis A':19,
'Hepatitis B':20,'Hepatitis C':21,'Hepatitis D':22,'Hepatitis E':23,'Alcoholic hepatitis':24,'Tuberculosis':25,
'Common Cold':26,'Pneumonia':27,'Dimorphic hemmorhoids(piles)':28,'Heart attack':29,'Varicose veins':30,'Hypothyroidism':31,
'Hyperthyroidism':32,'Hypoglycemia':33,'Osteoarthristis':34,'Arthritis':35,
'(vertigo) Paroymsal  Positional Vertigo':36,'Acne':37,'Urinary tract infection':38,'Psoriasis':39,
'Impetigo':40}},inplace=True)

X= df[l1]

y = df[["prognosis"]]
np.ravel(y)

def message():
    if (Symptom1.get() == "None" and  Symptom2.get() == "None" and Symptom3.get() == "None" and Symptom4.get() == "None" and Symptom5.get() == "None"):
        messagebox.showinfo("OPPS!!", "ENTER  SYMPTOMS PLEASE")
    else :
        NaiveBayes()
        SVM()
        RandomForest()
        DecisionTree()
        KNN()

# 1. NAIVE BAYES Method.
def NaiveBayes():
    from sklearn.naive_bayes import MultinomialNB
    gnb = MultinomialNB()
    gnb=gnb.fit(X,np.ravel(y))
    from sklearn.metrics import accuracy_score
    y_pred = gnb.predict(X_test)
    print("NAIVE BAYES Accuracy Score:", accuracy_score(y_test, y_pred)*100)
    # print(accuracy_score(y_test, y_pred, normalize=False))

    psymptoms = [Symptom1.get(),Symptom2.get(),Symptom3.get(),Symptom4.get(),Symptom5.get()]

    var1=Symptom1.get()
    var2=Symptom2.get()
    var3=Symptom3.get()
    var4=Symptom4.get()
    var5=Symptom5.get()
    
    input1 = var1
    input2 = var2
    input3 = var3
    input4 = var4
    input5 = var5

    # Check if the CSV file exists and is non-empty
    if os.path.exists(csv_file_path) and os.path.getsize(csv_file_path) > 0:
        # Append the new data to the existing CSV file
        existing_data = pd.read_csv(csv_file_path)
        new_data = pd.DataFrame({'Sympton1': [input1], 'Sympton2': [input2], 'Sympton3': [input3], 'Sympton4': [input4], 'Sympton5': [input5]})
        combined_data = pd.concat([existing_data, new_data], ignore_index=True)
        combined_data.to_csv(csv_file_path, index=False)
        # print("Data has been successfully added to", csv_file_path)
    else:
        # Create a new CSV file with the entered data
        new_data = pd.DataFrame({'Sympton1': [input1], 'Sympton2': [input2], 'Sympton3': [input3], 'Sympton4': [input4], 'Sympton5': [input5]})
        new_data.to_csv(csv_file_path, index=False)
        print("File", csv_file_path, "not found or empty. Created new file with the entered data.")

    for k in range(0,len(l1)):
        for z in psymptoms:
            if(z==l1[k]):
                l2[k]=1

    inputtest = [l2]
    predict = gnb.predict(inputtest)
    predicted=predict[0]

    h='no'
    for a in range(0,len(disease)):
        if(disease[predicted] == disease[a]):
            h='yes'
            break

    if (h=='yes'):
        t3.delete("1.0", END)
        t3.insert(END, disease[a])
        true_labels = y_test.values.ravel()
        calculate_display_metrics(true_labels, y_pred)
    else:
        t3.delete("1.0", END)
        t3.insert(END, "No Disease")
        

# 2. SVM Method.
def SVM():
    from sklearn.svm import SVC
    svm_model = SVC(kernel='linear')  
    svm_model.fit(X, np.ravel(y))

    from sklearn.metrics import accuracy_score
    y_pred = svm_model.predict(X_test)
    print("SVM Accuracy Score:", accuracy_score(y_test, y_pred)*100)
    # print(accuracy_score(y_test, y_pred, normalize=False))

    psymptoms = [Symptom1.get(), Symptom2.get(), Symptom3.get(), Symptom4.get(), Symptom5.get()]

    for k in range(0, len(l1)):
        for z in psymptoms:
            if z == l1[k]:
                l2[k] = 1

    inputtest = [l2]
    predict = svm_model.predict(inputtest)
    predicted = predict[0]

    h = 'no'
    for a in range(0, len(disease)):
        if disease[predicted] == disease[a]:
            h = 'yes'
            break

    if h == 'yes':
        t3.delete("1.0", END)
        t3.insert(END, disease[a])
        true_labels = y_test.values.ravel()
        calculate_display_metrics(true_labels, y_pred)
    else:
        t3.delete("1.0", END)
        t3.insert(END, "No Disease")


# 3. Random Forest

def RandomForest():
    clf = RandomForestClassifier()
    clf.fit(X, np.ravel(y))
    
    from sklearn.metrics import accuracy_score
    y_pred = clf.predict(X_test)
    print("Random Forest Accuracy Score:", accuracy_score(y_test, y_pred)*100)
    # print(accuracy_score(y_test, y_pred, normalize=False))
    
    psymptoms = [Symptom1.get(), Symptom2.get(), Symptom3.get(), Symptom4.get(), Symptom5.get()]
    for k in range(0, len(l1)):
        for z in psymptoms:
            if z == l1[k]:
                l2[k] = 1

    inputtest = [l2]
    predict = clf.predict(inputtest)
    predicted = predict[0]
    
    h = 'no'
    for a in range(0, len(disease)):
        if disease[predicted] == disease[a]:
            h = 'yes'
            break

    if h == 'yes':
        t3.delete("1.0", END)
        t3.insert(END, disease[a])
        true_labels = y_test.values.ravel()
        calculate_display_metrics(true_labels, y_pred)
    else:
        t3.delete("1.0", END)
        t3.insert(END, "No Disease")
        
# 4. Decision Tree     

def DecisionTree():
    clf = DecisionTreeClassifier()
    clf.fit(X, np.ravel(y))
    
    from sklearn.metrics import accuracy_score
    y_pred = clf.predict(X_test)
    print("Decision Tree Accuracy Score:", accuracy_score(y_test, y_pred)*100)
    # print(accuracy_score(y_test, y_pred, normalize=False))
    
    psymptoms = [Symptom1.get(), Symptom2.get(), Symptom3.get(), Symptom4.get(), Symptom5.get()]
    for k in range(0, len(l1)):
        for z in psymptoms:
            if z == l1[k]:
                l2[k] = 1

    inputtest = [l2]
    predict = clf.predict(inputtest)
    predicted = predict[0]
    
    h = 'no'
    for a in range(0, len(disease)):
        if disease[predicted] == disease[a]:
            h = 'yes'
            break

    if h == 'yes':
        t3.delete("1.0", END)
        t3.insert(END, disease[a])
        true_labels = y_test.values.ravel()
        calculate_display_metrics(true_labels, y_pred)
    else:
        t3.delete("1.0", END)
        t3.insert(END, "No Disease")

# 5. KNN

def KNN():
    clf = KNeighborsClassifier(n_neighbors=5)
    clf.fit(X, np.ravel(y))
    
    from sklearn.metrics import accuracy_score
    y_pred = clf.predict(X_test)
    print("KNN Accuracy Score:", accuracy_score(y_test, y_pred)*100)
    # print(accuracy_score(y_test, y_pred, normalize=False))

    psymptoms = [Symptom1.get(), Symptom2.get(), Symptom3.get(), Symptom4.get(), Symptom5.get()]
    for k in range(0, len(l1)):
        for z in psymptoms:
            if z == l1[k]:
                l2[k] = 1

    inputtest = [l2]
    predict = clf.predict(inputtest)
    predicted = predict[0]
    
    h = 'no'
    for a in range(0, len(disease)):
        if disease[predicted] == disease[a]:
            h = 'yes'
            break

    if h == 'yes':
        t3.delete("1.0", END)
        t3.insert(END, disease[a])
        true_labels = y_test.values.ravel()
        calculate_display_metrics(true_labels, y_pred)
    else:
        t3.delete("1.0", END)
        t3.insert(END, "No Disease")



def calculate_display_metrics(true_labels, predicted_labels):
    precision, recall, f1 = calculate_metrics(true_labels, predicted_labels)
    print("Precision:", precision*100)
    print("Recall:", recall*100)
    print("F1 Score:", f1*100)

# #.....
# # Function to plot the metrics
# def plot_metrics():
#     plt.figure(figsize=(10, 6))
#     plt.plot(algorithms, accuracy_scores, label='Accuracy', marker='o')
#     plt.plot(algorithms, precision_scores, label='Precision', marker='o')
#     plt.plot(algorithms, recall_scores, label='Recall', marker='o')
#     plt.plot(algorithms, f1_scores, label='F1 Score', marker='o')
#     plt.title('Performance Metrics by Algorithm')
#     plt.xlabel('Algorithm')
#     plt.ylabel('Score')
#     plt.legend()
#     plt.xticks(rotation=45)
#     plt.grid(True)
#     plt.tight_layout()
#     plt.show()

# Function to display message box
# def message():
#     if (Symptom1.get() == "None" and  Symptom2.get() == "None" and Symptom3.get() == "None" and Symptom4.get() == "None" and Symptom5.get() == "None"):
#         messagebox.showinfo("OPPS!!", "ENTER  SYMPTOMS PLEASE")
    # else :
        # calculate_plot_metrics()

##


root = Tk()
root.title(" Healthcare Revolution: Use of Machine Learning for Personalized Medicine and Disease Diagnosis")

root.configure()

# Prompt the user for input

Symptom1 = StringVar()
Symptom1.set(None)
Symptom2 = StringVar()
Symptom2.set(None)
Symptom3 = StringVar()
Symptom3.set(None)
Symptom4 = StringVar()
Symptom4.set(None)
Symptom5 = StringVar()
Symptom5.set(None)

w2 = Label(root, justify=LEFT, text=" Disease Prediction From Symptoms ")
w2.config(font=("Elephant", 30))
w2.grid(row=1, column=0, columnspan=3, padx=100)

NameLb1 = Label(root, text="")
NameLb1.config(font=("Elephant", 20))
NameLb1.grid(row=5, column=1, pady=10,  sticky=W)

S1Lb = Label(root,  text="Symptom 1")
S1Lb.config(font=("Elephant", 15))
S1Lb.grid(row=7, column=1, pady=10 , sticky=W)

S2Lb = Label(root,  text="Symptom 2")
S2Lb.config(font=("Elephant", 15))
S2Lb.grid(row=8, column=1, pady=10, sticky=W)

S3Lb = Label(root,  text="Symptom 3")
S3Lb.config(font=("Elephant", 15))
S3Lb.grid(row=9, column=1, pady=10, sticky=W)

S4Lb = Label(root,  text="Symptom 4")
S4Lb.config(font=("Elephant", 15))
S4Lb.grid(row=10, column=1, pady=10, sticky=W)

S5Lb = Label(root,  text="Symptom 5")
S5Lb.config(font=("Elephant", 15))
S5Lb.grid(row=11, column=1, pady=10, sticky=W)

lr = Button(root, text="Predict",height=2, width=20, command=message)
lr.config(font=("Elephant", 15))
lr.grid(row=15, column=1,pady=20)

OPTIONS = sorted(l1)

S1En = OptionMenu(root, Symptom1,*OPTIONS)
S1En.grid(row=7, column=2)

S2En = OptionMenu(root, Symptom2,*OPTIONS)
S2En.grid(row=8, column=2)

S3En = OptionMenu(root, Symptom3,*OPTIONS)
S3En.grid(row=9, column=2)

S4En = OptionMenu(root, Symptom4,*OPTIONS)
S4En.grid(row=10, column=2)

S5En = OptionMenu(root, Symptom5,*OPTIONS)
S5En.grid(row=11, column=2)

NameLb = Label(root, text="")
NameLb.config(font=("Elephant", 20))
NameLb.grid(row=13, column=1, pady=10,  sticky=W)

NameLb = Label(root, text="")
NameLb.config(font=("Elephant", 15))
NameLb.grid(row=18, column=1, pady=10,  sticky=W)

t3 = Text(root, height=2, width=30)
t3.config(font=("Elephant", 20))
t3.grid(row=20, column=1 , padx=10)



root.mainloop()